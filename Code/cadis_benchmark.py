# -*- coding: utf-8 -*-
"""
Benchmark Pgmpy cadis algos on the 'Asia' dataset.
Outputs results (algorithm, parameters, TP, TN, precision, recall, F1, SHD)
to an Excel spreadsheet.

Created on Thu Mar  5 11:28:38 2026
@author: Saptarshi Pyne
"""

## The 'Asia' dataset is taken from:
## Fig. 2, Lauritzen, Steffen L., and David J. Spiegelhalter.
## "Local computations with probabilities on graphical structures and
## their application to expert systems." Journal of the Royal
## Statistical Society: Series B (Methodological) 50.2 (1988): 157-194.

######################################################
## Begin: Imports
######################################################

import os
from pgmpy.utils import get_example_model
import pgmpy.estimators

import numpy as np
import networkx as nx
import sklearn.metrics

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

######################################################
## End: Imports
######################################################

## Load the Asia network
asia_model = get_example_model('asia')

## Simulate 1,000 samples
asia_data = asia_model.simulate(n_samples=int(1e3), seed=42)


######################################################
## Begin: Metric Computation
######################################################

def compute_metrics(estimated_model, true_model):
    """
    Compute TP, TN, Precision, Recall, F1-score, and SHD
    between the skeleton of the estimated and true DAGs.

    Parameters
    ----------
    estimated_model : pgmpy DAG / PDAG / BayesianNetwork
    true_model      : pgmpy BayesianNetwork (ground truth)

    Returns
    -------
    dict with keys: TP, TN, Precision, Recall, F1, SHD
    """
    nodes = list(true_model.nodes())

    # Use the node ordering from the true model so matrices align
    est_adj = nx.to_numpy_array(
        estimated_model.to_undirected(), nodelist=nodes, weight=None
    )
    true_adj = nx.to_numpy_array(
        true_model.to_undirected(), nodelist=nodes, weight=None
    )

    y_true = np.ravel(true_adj).astype(int)
    y_pred = np.ravel(est_adj).astype(int)

    tn, fp, fn, tp = sklearn.metrics.confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    precision = sklearn.metrics.precision_score(y_true, y_pred, zero_division=0)
    recall    = sklearn.metrics.recall_score(y_true, y_pred, zero_division=0)
    f1        = sklearn.metrics.f1_score(y_true, y_pred, zero_division=0)

    # Structural Hamming Distance (SHD) on skeletons:
    # count of edges that differ between true and estimated undirected graphs
    true_edges = set(frozenset(e) for e in true_model.to_undirected().edges())
    est_edges  = set(frozenset(e) for e in estimated_model.to_undirected().edges())
    shd = len(true_edges.symmetric_difference(est_edges))

    return {
        "TP": int(tp),
        "TN": int(tn),
        "Precision": round(float(precision), 4),
        "Recall": round(float(recall), 4),
        "F1": round(float(f1), 4),
        "SHD": int(shd),
    }


######################################################
## End: Metric Computation
######################################################


######################################################
## Begin: Run Algorithms
######################################################

results = []  # list of dicts: {Algorithm, Parameters, TP, TN, Precision, Recall, F1, SHD}


def run_and_record(algo_name, params_dict, model):
    metrics = compute_metrics(model, asia_model)
    row = {"Algorithm": algo_name, "Parameters": str(params_dict)}
    row.update(metrics)
    results.append(row)
    print(f"[DONE] {algo_name} | {params_dict} | F1={metrics['F1']} SHD={metrics['SHD']}")


# ------------------------------------------------------------------
# 1. PC Algorithm
#    Ref: Koller, D., & Friedman, N. (2009). Probabilistic Graphical Models.
#         Neapolitan, R.E. (2004). Learning Bayesian Networks.
#    pgmpy docs: https://pgmpy.org/structure_estimator/pc.html
# ------------------------------------------------------------------
for ci_test in ['chi_square', 'g_sq']:
    for variant in ['stable', 'orig', 'parallel']:
        for max_cond_vars in [2, 4]:
            params = {
                "ci_test": ci_test,
                "variant": variant,
                "max_cond_vars": max_cond_vars,
            }
            try:
                est = pgmpy.estimators.PC(data=asia_data)
                model = est.estimate(
                    ci_test=ci_test,
                    variant=variant,
                    max_cond_vars=max_cond_vars,
                    return_type='dag',
                    show_progress=False,
                )
                run_and_record("PC", params, model)
            except Exception as e:
                print(f"[SKIP] PC | {params} | Error: {e}")


# ------------------------------------------------------------------
# 2. HillClimbSearch (Greedy Hill Climbing)
#    Ref: Russell, S., & Norvig, P. (2009). Artificial Intelligence: A Modern Approach.
#         Chickering, D.M. (2002). Optimal Structure Identification With Greedy Search.
#    pgmpy docs: https://pgmpy.org/structure_estimator/hill_climb_search.html
# ------------------------------------------------------------------
## Scoring method string names as used by the current pgmpy API:
## 'k2' = K2Score, 'bdeu' = BDeuScore, 'bic-d' = BicScore (discrete)
## Ref: https://pgmpy.org/structure_estimator/hill.html
SCORING_METHODS = {'K2Score': 'k2', 'BDeuScore': 'bdeu', 'BicScore': 'bic-d'}

for scoring_method_name in ['K2Score', 'BDeuScore', 'BicScore']:
    for max_indegree in [2, 4]:
        params = {
            "scoring_method": scoring_method_name,
            "max_indegree": max_indegree,
        }
        try:
            est = pgmpy.estimators.HillClimbSearch(data=asia_data)
            model = est.estimate(
                scoring_method=SCORING_METHODS[scoring_method_name],
                max_indegree=max_indegree,
                show_progress=False,
            )
            run_and_record("HillClimbSearch", params, model)
        except Exception as e:
            print(f"[SKIP] HillClimbSearch | {params} | Error: {e}")


# ------------------------------------------------------------------
# 3. ExhaustiveSearch
#    Ref: Koller & Friedman (2009). Probabilistic Graphical Models.
#    pgmpy docs: https://pgmpy.org/structure_estimator/exhaustive.html
#    Note: Only feasible for small networks; Asia has 8 nodes so we limit
#          to a 5-node subset to keep runtime manageable.
# ------------------------------------------------------------------
subset_nodes = ['asia', 'tub', 'smoke', 'lung', 'bronc']
asia_data_subset = asia_data[subset_nodes]

for scoring_method_name in ['K2Score', 'BDeuScore', 'BicScore']:
    params = {
        "scoring_method": scoring_method_name,
        "node_subset": subset_nodes,
    }
    try:
        est = pgmpy.estimators.ExhaustiveSearch(data=asia_data_subset)
        model = est.estimate(scoring_method=SCORING_METHODS[scoring_method_name])
        run_and_record("ExhaustiveSearch", params, model)
    except Exception as e:
        print(f"[SKIP] ExhaustiveSearch | {params} | Error: {e}")



# ------------------------------------------------------------------
# 4. GES (Greedy Equivalence Search)
#    Ref: Chickering, D.M. (2002). Optimal structure identification with
#         greedy search. Journal of Machine Learning Research, 3, 507-554.
#    pgmpy docs: https://pgmpy.org/structure_estimator/ges.html
#
#    Works in 3 phases over equivalence classes (PDAGs):
#      (1) Forward: add edges while score improves
#      (2) Backward: remove edges while score improves
#      (3) Turning: flip orientations while score improves
#    Supports same scoring strings as HillClimbSearch, plus 'aic-d'.
# ------------------------------------------------------------------
GES_SCORING_METHODS = ['k2', 'bdeu', 'bic-d', 'aic-d']

for scoring_method in GES_SCORING_METHODS:
    params = {"scoring_method": scoring_method}
    try:
        est = pgmpy.estimators.GES(data=asia_data)
        model = est.estimate(scoring_method=scoring_method)
        run_and_record("GES", params, model)
    except Exception as e:
        print(f"[SKIP] GES | {params} | Error: {e}")


# ------------------------------------------------------------------
# 5. MmhcEstimator (Max-Min Hill Climbing)
#    Ref: Tsamardinos, I., Brown, L.E., & Aliferis, C.F. (2006).
#         The max-min hill-climbing Bayesian network structure learning
#         algorithm. Machine Learning, 65(1), 31-78.
#    pgmpy docs: https://pgmpy.org/structure_estimator/mmhc.html
#
#    Hybrid algorithm:
#      Phase 1 (constraint-based): MMPC finds candidate parent/child sets
#      Phase 2 (score-based): Hill Climbing orients edges within skeleton
#    Parameters:
#      significance_level: p-value threshold for MMPC independence tests
#      scoring_method: scoring for Hill Climbing orientation phase
#      tabu_length: prevents reversing the last N graph modifications
# ------------------------------------------------------------------
for scoring_method in ['k2', 'bdeu', 'bic-d']:
    for significance_level in [0.01, 0.05]:
        for tabu_length in [10, 20]:
            params = {
                "scoring_method": scoring_method,
                "significance_level": significance_level,
                "tabu_length": tabu_length,
            }
            try:
                est = pgmpy.estimators.MmhcEstimator(data=asia_data)
                model = est.estimate(
                    scoring_method=scoring_method,
                    significance_level=significance_level,
                    tabu_length=tabu_length,
                )
                run_and_record("MmhcEstimator", params, model)
            except Exception as e:
                print(f"[SKIP] MmhcEstimator | {params} | Error: {e}")


# ------------------------------------------------------------------
# 6. TreeSearch — Chow-Liu
#    Ref: Chow, C.K., & Liu, C.N. (1968). Approximating discrete
#         probability distributions with dependence trees.
#         IEEE Transactions on Information Theory, 14(3), 462-467.
#    pgmpy docs: https://pgmpy.org/structure_estimator/tree.html
#
#    Finds the maximum-weight spanning tree using mutual information
#    as edge weights, then orients edges away from the root node.
#    Note: TreeSearch produces a tree (each node has at most 1 parent),
#    so it is a restricted structural class compared to a general DAG.
#    We vary root_node across all 8 Asia variables.
# ------------------------------------------------------------------
asia_nodes = list(asia_model.nodes())

for root_node in asia_nodes:
    params = {
        "estimator_type": "chow-liu",
        "root_node": root_node,
    }
    try:
        est = pgmpy.estimators.TreeSearch(data=asia_data, root_node=root_node)
        model = est.estimate(estimator_type='chow-liu', show_progress=False)
        run_and_record("TreeSearch (Chow-Liu)", params, model)
    except Exception as e:
        print(f"[SKIP] TreeSearch (Chow-Liu) | {params} | Error: {e}")


# ------------------------------------------------------------------
# 7. TreeSearch — TAN (Tree-Augmented Naive Bayes)
#    Ref: Friedman, N., Geiger, D., & Goldszmidt, M. (1997).
#         Bayesian network classifiers. Machine Learning, 29, 131-163.
#    pgmpy docs: https://pgmpy.org/structure_estimator/tree.html
#
#    Extends Chow-Liu for classification: adds edges from a designated
#    class_node to every feature node, then fits a tree over features.
#    class_node: the target variable to classify (dysp = shortness of
#      breath, lung = lung cancer — the two most clinically meaningful
#      outcomes in the Asia network).
#    root_node=None: auto-picks the node with the highest sum of edge
#      weights, removing the need to loop over root choices.
# ------------------------------------------------------------------
TAN_CLASS_NODES = ['dysp', 'lung']

for class_node in TAN_CLASS_NODES:
    params = {
        "estimator_type": "tan",
        "class_node": class_node,
        "root_node": "auto",
    }
    try:
        est = pgmpy.estimators.TreeSearch(data=asia_data, root_node=None)
        model = est.estimate(
            estimator_type='tan',
            class_node=class_node,
            show_progress=False,
        )
        run_and_record("TreeSearch (TAN)", params, model)
    except Exception as e:
        print(f"[SKIP] TreeSearch (TAN) | {params} | Error: {e}")


######################################################
## End: Run Algorithms
######################################################


######################################################
## Begin: Write Excel Output
######################################################

COLUMNS = ["Algorithm", "Parameters", "F1", "TP", "TN", "Precision", "Recall", "SHD"]

wb = Workbook()
ws = wb.active
ws.title = "Cadis Benchmark – Asia"

# ---- Styles ----
header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill   = PatternFill("solid", start_color="2F4F8F")   # dark blue
alt_fill      = PatternFill("solid", start_color="EBF0FA")   # light blue
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="B0B8C8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---- Column widths ----
col_widths = {
    "A": 22,  # Algorithm
    "B": 52,  # Parameters
    "C": 10,  # F1
    "D": 8,   # TP
    "E": 8,   # TN
    "F": 12,  # Precision
    "G": 10,  # Recall
    "H": 8,   # SHD
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width
ws.row_dimensions[1].height = 28

# ---- Header row ----
for col_idx, col_name in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = border

# ---- Data rows ----
for row_idx, record in enumerate(results, start=2):
    fill = alt_fill if row_idx % 2 == 0 else PatternFill()  # zebra striping
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        value = record.get(col_name, "")
        cell  = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = fill
        cell.border    = border
        cell.alignment = left_align if col_name in ("Algorithm", "Parameters") else center_align

# Freeze header row
ws.freeze_panes = "A2"

output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cadis_benchmark_results.xlsx")
wb.save(output_path)
print(f"\nResults saved to: {output_path}")

######################################################
## End: Write Excel Output
######################################################
